#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import re
import io
import os
import openpyxl
import datetime
import time
import xlsxwriter
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from zipfile import ZipFile
import shutil
import subprocess
from openpyxl.styles import Font, PatternFill


PROJECT_HOME = os.environ['PROJECT_HOME']
print(PROJECT_HOME)


class REPORT():
    def __init__(self):

        self.master_array_wr = []
        self.master_array_default = []
        self.master_array = []
        self.master_array_merge = []
        self.current_time = ""
        self.repo_name = ""

    def main(self):

        # Print a message indicating that the script has entered 'report.py'
        print("Entered report.py")

        self.current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        outputpath = PROJECT_HOME + "/reports/" + f'{self.current_time}' + "/"
        if not os.path.exists(outputpath):
            os.makedirs(outputpath)

        fp = open(PROJECT_HOME + "/test_script", "r")
        lines = fp.readlines()
        self.exe_count = 0
        m_count = 0
        for i in lines:
            if re.search("^\./.*sh.*", i, re.IGNORECASE):
                # Format and get current time for filenames

                self.exe_count += 1
                val = i.split("/")
                self.repo_name = val[3]
                print(self.repo_name)

                self.master_array_wr = []
                self.master_array_default = []

                # Open a file for reading containing test execution commands
                file = open(PROJECT_HOME + "/tests/" + self.repo_name  + "/" + self.repo_name + "_execution_18a.sh", "r")
                content = file.read().splitlines()

                count_d = 0
                count_w = 0
                count_g = 0

                # Loop through each line of the content
                for i in range(0, len(content)):

                    if re.match("^cd.*", content[i]):
                        path = content[i].split(" ")[1].replace('$PROJECT_HOME', str(PROJECT_HOME))
                        testname = str(content[i].split(" ")[1].split("/")[-1])
                        access_name = str(content[i].split(" ")[1].split("/")[-2])
                        a = []
                        b = []
                        c = []
                        d = []

                        # Open and read the test log file for write_read_expect tests
                        time.sleep(4)
                        fp1 = open(path + "/" + testname + ".log", "r")
                        word = fp1.readlines()
                        # Open and read the test log file for register_default_read test
                        # Check test status based on log content and test type
                        m_count = m_count + 1
                        if re.match(".*test pass.*", word[-1]):
                            # Get log path for write_read_expect tests
                            status = "PASSED"
                        else:
                            status = "FAILED"

                        logpath = path + "/"+testname + ".log"

                        if "default" in access_name or "initial" in access_name:
                            count_d = count_d + 1
                            a.append(count_d)
                            a.append(testname)
                            a.append(status)
                            a.append(logpath)
                        elif "write" in access_name:
                            count_w = count_w + 1
                            b.append(count_w)
                            b.append(testname)
                            b.append(status)
                            b.append(logpath)
                        else:
                            count_g = count_g + 1
                            d.append(count_g)
                            d.append(testname)
                            d.append(status)
                            d.append(logpath)

                        c.append(m_count)
                        c.append(testname)
                        c.append(status)
                        c.append(logpath)

                        self.master_array_default.append(a)
                        self.master_array_wr.append(b)
                        self.master_array_merge.append(c)
                        self.master_array.append(d)

                if self.exe_count >= 2:
                    df = pd.DataFrame(self.master_array_merge, columns=['count', 'Testname', 'Status', 'Logpath'])

                    # Write the DataFrame to an Excel file with a timestamp-based filename
                    df.to_excel(outputpath + f'{self.current_time}_merge_output.xlsx', sheet_name='sheet_name',
                                index=False)

                    # Load the Excel file
                    workbook = openpyxl.load_workbook(
                        outputpath + f'{self.current_time}_merge_output.xlsx')  # Replace with your Excel file name
                    worksheet = workbook.active
                    self.color(workbook)
                    self.pie_chart(workbook)
                    # Save the modified Excel file
                    workbook.save(outputpath + f'{self.current_time}_merge_output.xlsx')

                # Create a DataFrame from the master_array with specified column names
                if any(self.master_array_wr):
                    df1 = pd.DataFrame(self.master_array_wr, columns=['count', 'Testname', 'Status', 'Logpath'])
                    df1.to_excel(outputpath + f'{self.repo_name}_{self.current_time}test_output_wr.xlsx',
                                 sheet_name='sheet_name', index=False)
                    workbook1 = openpyxl.load_workbook(
                        outputpath + f'{self.repo_name}_{self.current_time}test_output_wr.xlsx')
                    self.color(workbook1)
                    self.pie_chart(workbook1)
                    workbook1.save(outputpath + f'{self.repo_name}_{self.current_time}test_output_wr.xlsx')


                if any(self.master_array_default):
                    df2 = pd.DataFrame(self.master_array_default, columns=['count', 'Testname', 'Status', 'Logpath'])
                    df2.to_excel(outputpath + f'{self.repo_name}_{self.current_time}test_output_default.xlsx',
                                 sheet_name='sheet_name', index=False)
                    workbook2 = openpyxl.load_workbook(
                        outputpath + f'{self.repo_name}_{self.current_time}test_output_default.xlsx')
                    self.color(workbook2)
                    self.pie_chart(workbook2)
                    workbook2.save(outputpath + f'{self.repo_name}_{self.current_time}test_output_default.xlsx')

                if any(self.master_array):
                    df3 = pd.DataFrame(self.master_array, columns=['count', 'Testname', 'Status', 'Logpath'])
                    df3.to_excel(outputpath + f'{self.repo_name}_{self.current_time}test_output.xlsx',
                                 sheet_name='sheet_name', index=False)
                    workbook3 = openpyxl.load_workbook(
                        outputpath + f'{self.repo_name}_{self.current_time}test_output.xlsx')
                    self.color(workbook3)
                    self.pie_chart(workbook3)
                    workbook3.save(outputpath + f'{self.repo_name}_{self.current_time}test_output.xlsx')

    def pie_chart(self, workbook):
        # Load the Excel file
        # Replace with your Excel file name
        sheet = workbook.active

        # Read data from columns
        test_counts = [cell.value for cell in sheet['A'][1:]]
        test_statuses = [cell.value for cell in sheet['C'][1:]]

        # Count the number of tests and their results
        # total_tests = len(test_counts)
        tests_passed = test_statuses.count('PASSED')
        tests_failed = test_statuses.count('FAILED')
        total_tests = tests_passed + tests_failed

        # Create a pie chart using matplotlib
        labels = ['PASSED', 'FAILED']
        sizes = [tests_passed, tests_failed]
        colors = ['lightgreen', 'red']

        # Create a string with the counts for the labels
        label_counts = [f'{label}\n{size}' for label, size in zip(labels, sizes)]

        plt.pie(sizes, labels=label_counts, colors=colors, autopct='', startangle=140)
        plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

        # Add the total test count outside the pie chart
        total_label = f'Total: {total_tests}'
        plt.text(0.5, 1.05, total_label, horizontalalignment='center', verticalalignment='center',
                 transform=plt.gca().transAxes)

        plt.savefig('pie_chart.png')  # Save the pie chart as an image
        plt.close()  # Close the matplotlib figure to free up resources

        # Add the image to the worksheet
        img = Image('pie_chart.png')  # Replace 'your_image.png' with the actual image file path
        img.anchor = 'G2'  # Set the cell where the image will be place	d
        sheet.add_image(img)

        # Function to zip all log files into a single ZIP file
    def zip_logs(self,input_excel, output_path, output_file_name):
        try:
            # Read the input Excel file using 'openpyxl' engine
            df = pd.read_excel(input_excel, engine='openpyxl')

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return

        log_files = []  # List to store log file paths

        # Iterate through rows in the DataFrame
        for index, row in df.iterrows():
            log_file_path = row['Logpath']
            print(f"Checking log file: {log_file_path}")
            # Check if the file exists and has a '.log' extension
           # if os.path.exists(log_file_path) and log_file_path.endswith('.log'):
            log_files.append(log_file_path)


        if log_files:
           # Create a ZIP file in the specified output path
           with ZipFile(os.path.join(output_path, output_file_name), 'w') as zipf:
               for log_file_path in log_files:
                   log_file_name = os.path.basename(log_file_path)
                   # Add each log file to the ZIP file with its original name
                   zipf.write(log_file_path, log_file_name)

    def color(self, workbook):
        worksheet = workbook.active
        fill = PatternFill(start_color="99CCFF", end_color="99CCFF",fill_type="solid")  # Iterate through the first row (row 1) and apply the fill style to each cell

        for cell in worksheet[1]:
            cell.fill = fill
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                if cell.value == 'FAILED':
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
                elif cell.value == 'PASSED':
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green



if __name__ == "__main__":
    obj = REPORT()
    obj.main()

    output_directory = PROJECT_HOME + "/reports/" + f'{obj.current_time}' + "/"
    output_directory1 = PROJECT_HOME + "/logs/" + f'{obj.current_time}'
    output_file_name = f'all_logs_{obj.current_time}.zip'  # Change this to your desired ZIP file name
    if not os.path.exists(output_directory1):
        os.makedirs(output_directory1)
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    if obj.exe_count >=2:
        excel_file_path = os.path.join(output_directory, f'{obj.current_time}_merge_output.xlsx')
        obj.zip_logs(excel_file_path, output_directory, output_file_name)
        shutil.move(os.path.join(output_directory, output_file_name), output_directory1)


