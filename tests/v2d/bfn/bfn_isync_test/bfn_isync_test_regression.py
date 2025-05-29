#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("bfn_isync_test_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
fail_cnt = 0

for i in range(2,row_count +1):
    v2d_num = wb1.cell(row=i,column=1).value
    drf_clock = wb1.cell(row=i,column=2).value
    rv = subprocess.call(['python bfn_isync_test.py '  f'{v2d_num} '  f'{drf_clock} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
