#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
wb = openpyxl.Workbook()

wa = wb.active
header = ["v2d_id" ,"drf_clock"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
row += 1


for v2d_num in range (4):
    for drf_clock in range(1):
        wa.cell(row,1).value = v2d_num
        wa.cell(row,2).value = drf_clock
        row+=1

sheet_name= wb['Sheet']
sheet_name.title ='bfn_test_list'

wb.save("bfn_test_list.xlsx")

