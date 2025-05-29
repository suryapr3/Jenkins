#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("post_proessin_2_channel_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0

for i in range(2,row_count +1):
    quad_instance = wb1.cell(row=i,column=1).value
    apb_num = wb1.cell(row=i,column=2).value
    drf_clock = wb1.cell(row=i,column=3).value
    test_config = wb1.cell(row=i,column=4).value
    delay_value = wb1.cell(row=i,column=5).value
    subprocess.call(['python post_processing_2_channel.py '  f'{quad_instance} '  f'{apb_num} ' f'{drf_clock} ' f'{test_config} ' f'{delay_value} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
