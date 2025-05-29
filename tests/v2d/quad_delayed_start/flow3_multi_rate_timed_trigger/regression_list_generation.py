#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

V2D = fw.ABC_V2D()
wa = wb.active
header = ["quad_instance", "drf_clock", "test_config", "filter_mode", "delay_value"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

V2D.drf_clock_check(drf_clock)
print("all parameters are proper\n")
test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

for quad_instance in range(2):
    for test_config in test_config_list:
        for filter_mode in range(2):
            wa.cell(row,1).value = quad_instance
            wa.cell(row,2).value = drf_clock
            wa.cell(row,3).value = test_config
            wa.cell(row,4).value = filter_mode
            wa.cell(row,5).value = 768  #delay value
            row+=1

sheet_name= wb['Sheet']
sheet_name.title ='flow3_multi_rate_timed_trigger'

wb.save("flow3_multi_rate_timed_trigger_list.xlsx")

