#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("post_processing_2streams_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
fail_cnt = 0

for i in range(2,row_count +1):
    vex_tx_b = wb1.cell(row=i,column=1).value
    vex_rx_d0 = wb1.cell(row=i,column=2).value
    vex_rx_d1 = wb1.cell(row=i,column=3).value
    vex_rx_f = wb1.cell(row=i,column=4).value
    v2d_num = wb1.cell(row=i,column=5).value
    apb_num = wb1.cell(row=i,column=6).value
    drf_clock = wb1.cell(row=i,column=7).value
    test_config = wb1.cell(row=i,column=8).value
    rv = subprocess.call(['python post_processing_2streams.py '  f'{vex_tx_b} '  f'{vex_rx_d0} ' f'{vex_rx_d1} ' f'{vex_rx_f} ' f'{v2d_num} '  f'{apb_num} '  f'{drf_clock} ' f'{test_config} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
