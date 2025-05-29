#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("tx_hbi_0_nocfg_drop_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0

for i in range(2,row_count +1):
    tx_vex_num = wb1.cell(row=i,column=1).value
    rx_vex_num = wb1.cell(row=i,column=2).value
    v2d_num = wb1.cell(row=i,column=3).value
    apb_num = wb1.cell(row=i,column=4).value
    drf_clock = wb1.cell(row=i,column=5).value
    sys_rate = wb1.cell(row=i,column=6).value
    subprocess.call(['python tx_hbi_0_nocfg_drop.py '  f'{tx_vex_num} '  f'{rx_vex_num} ' f'{v2d_num} '  f'{apb_num} '  f'{drf_clock} ' f'{sys_rate} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
