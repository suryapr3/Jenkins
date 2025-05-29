#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
wb = openpyxl.Workbook()

wa = wb.active
header = ["v2d_id" ,"apb_id" ,"drf_clock" ,"sys_rate"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
row += 1

sys_rate_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 21, 22, 23, 24, 25, 26]

for v2d_num in range(4):
    for apb_num in range (2):
        for drf_clock in range(1):
            for sys_rate in sys_rate_list:
                if(drf_clock == 0) and (sys_rate == 11):
                    continue
                wa.cell(row,1).value = v2d_num
                wa.cell(row,2).value = apb_num
                wa.cell(row,3).value = drf_clock
                wa.cell(row,4).value = sys_rate
                row+=1

sheet_name= wb['Sheet']
sheet_name.title ='tx_no_data_start'

wb.save("tx_no_data_start_list.xlsx")

