#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("buffer_b_c_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
fail_cnt = 0

for i in range(2,row_count +1):
    tx1_vex_num = wb1.cell(row=i,column=1).value
    tx2_vex_num = wb1.cell(row=i,column=2).value
    rx_vex_num = wb1.cell(row=i,column=3).value
    v2d_num = wb1.cell(row=i,column=4).value
    apb_num = wb1.cell(row=i,column=5).value
    drf_clock = wb1.cell(row=i,column=6).value
    test_config = wb1.cell(row=i,column=7).value
    rv = subprocess.call(['python buffer_b_c.py '  f'{tx1_vex_num} '  f'{tx2_vex_num} ' f'{rx_vex_num} ' f'{v2d_num} '  f'{apb_num} '  f'{drf_clock} ' f'{test_config} '], shell=True)
    if (1 == rv):
        fail_cnt = fail_cnt + 1;
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(10/1000)

print("regression completed\n");
