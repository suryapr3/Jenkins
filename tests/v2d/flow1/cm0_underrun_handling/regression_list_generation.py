#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
wb = openpyxl.Workbook()

wa = wb.active
header = ["tx_vex_id", "rx_vex_id", "v2d_id", "apb_id", "drf_clock", "sys_rate"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
wa.cell(row, 6).value = header[5]
row += 1

v2d_0_valid_vex_list = [0, 1, 8, 9, 32]
v2d_1_valid_vex_list = [2, 3, 10, 11, 34]
v2d_2_valid_vex_list = [4, 5, 12, 13, 36]
v2d_3_valid_vex_list = [6, 7, 14, 15, 38]
sys_rate_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]

for tx_vex_num in v2d_0_valid_vex_list:
    for rx_vex_num in v2d_0_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for drf_clock in range(1):
                for sys_rate in sys_rate_list:
                    if (drf_clock == 0 and sys_rate == 11):
                        continue
                    wa.cell(row,1).value = tx_vex_num
                    wa.cell(row,2).value = rx_vex_num
                    wa.cell(row,3).value = 0
                    wa.cell(row,4).value = apb_num
                    wa.cell(row,5).value = drf_clock
                    wa.cell(row,6).value = sys_rate
                    row+=1

for tx_vex_num in v2d_1_valid_vex_list:
    for rx_vex_num in v2d_1_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for drf_clock in range(1):
                for sys_rate in sys_rate_list:
                    if (drf_clock == 0 and sys_rate == 11):
                        continue
                    wa.cell(row,1).value = tx_vex_num
                    wa.cell(row,2).value = rx_vex_num
                    wa.cell(row,3).value = 1
                    wa.cell(row,4).value = apb_num
                    wa.cell(row,5).value = drf_clock
                    wa.cell(row,6).value = sys_rate
                    row+=1

for tx_vex_num in v2d_2_valid_vex_list:
    for rx_vex_num in v2d_2_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for drf_clock in range(1):
                for sys_rate in sys_rate_list:
                    if (drf_clock == 0 and sys_rate == 11):
                        continue
                    wa.cell(row,1).value = tx_vex_num
                    wa.cell(row,2).value = rx_vex_num
                    wa.cell(row,3).value = 2
                    wa.cell(row,4).value = apb_num
                    wa.cell(row,5).value = drf_clock
                    wa.cell(row,6).value = sys_rate
                    row+=1

for tx_vex_num in v2d_3_valid_vex_list:
    for rx_vex_num in v2d_3_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for drf_clock in range(1):
                for sys_rate in sys_rate_list:
                    if (drf_clock == 0 and sys_rate == 11):
                        continue
                    wa.cell(row,1).value = tx_vex_num
                    wa.cell(row,2).value = rx_vex_num
                    wa.cell(row,3).value = 3
                    wa.cell(row,4).value = apb_num
                    wa.cell(row,5).value = drf_clock
                    wa.cell(row,6).value = sys_rate
                    row+=1


#for vex_num in v2d_1_valid_vex_list:
#        for apb_num in range(2):
#                for sys_rate in sys_rate_list:
#                        wa.cell(row, 1).value = vex_num
#                        wa.cell(row, 2).value = 1
#                        wa.cell(row, 3).value = apb_num
#                        wa.cell(row, 4).value = sys_rate
#                        row+=1
#
#for vex_num in v2d_2_valid_vex_list:
#        for apb_num in range(2):
#                for sys_rate in sys_rate_list:
#                        wa.cell(row, 1).value = vex_num
#                        wa.cell(row, 2).value = 2
#                        wa.cell(row, 3).value = apb_num
#                        wa.cell(row, 4).value = sys_rate
#                        row+=1
#
#for vex_num in v2d_3_valid_vex_list:
#        for apb_num in range(2):
#                for sys_rate in sys_rate_list:
#                        wa.cell(row, 1).value = vex_num
#                        wa.cell(row, 2).value = 3
#                        wa.cell(row, 3).value = apb_num
#                        wa.cell(row, 4).value = sys_rate
#                        row+=1

sheet_name= wb['Sheet']
sheet_name.title ='cm0_underrun_handling'

wb.save("cm0_underrun_handling.xlsx")

