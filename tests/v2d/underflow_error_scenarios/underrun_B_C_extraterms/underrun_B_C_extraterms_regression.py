#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("underrun_B_C_extraterms_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
fail_cnt = 0

for i in range(2,row_count +1):
    vex_num0 = wb1.cell(row=i,column=1).value
    vex_num1 = wb1.cell(row=i,column=2).value
    v2d_num = wb1.cell(row=i,column=3).value
    apb_num = wb1.cell(row=i,column=4).value
    pdpd_rate = wb1.cell(row=i,column=5).value
    rv = subprocess.call(['python underrun_B_C_extraterms.py '  f'{vex_num0} '  f'{vex_num1} ' f'{v2d_num} '  f'{apb_num} ' f'{pdpd_rate} '], shell=True)
    regression_count = regression_count + 1
    if (1 == rv):
        fail_cnt = fail_cnt + 1;
    print(f"TC completed:{regression_count}\n")
    time.sleep(10/1000)

print("regression completed\n");
if (fail_cnt != 0):
    print(f"number of failure iteration:{fail_cnt}\n")
else:
    print("All iterations passed\n")
