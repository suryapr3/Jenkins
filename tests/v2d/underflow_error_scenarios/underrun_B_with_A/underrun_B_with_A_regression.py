#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("underrun_B_with_A_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
error_count = 0

for i in range(2,row_count +1):
    vex_num = wb1.cell(row=i,column=1).value
    v2d_num = wb1.cell(row=i,column=2).value
    apb_num = wb1.cell(row=i,column=3).value
    pdpd_rate = wb1.cell(row=i,column=4).value
    rv = subprocess.call(['python underrun_B_with_A.py '  f'{vex_num} '  f'{v2d_num} '  f'{apb_num} '  f'{pdpd_rate} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    if not (rv == 0):
        error_count = error_count + 1;
    time.sleep(100/1000)

print("regression completed\n");
if error_count == 0:
    print("all iterations are passed\n")
else:
    print(f"{error_count} iterations are faild\n")
