#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
wb = openpyxl.Workbook()

wa = wb.active
header = ["vex_id0", "vex_id1", "v2d_id", "apb_id", "pdpd_rate"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
row += 1

v2d_0_valid_vex_list = [0, 8, 16, 24, 32]
v2d_0_valid_vex1_list = [0, 1, 8, 9, 16, 17, 24, 25, 32, 33]
v2d_1_valid_vex_list = [2, 10, 34]
v2d_1_valid_vex1_list = [2, 3, 10, 11, 34, 35]
v2d_2_valid_vex_list = [4, 12, 36]
v2d_2_valid_vex1_list = [4, 5, 12, 13, 36, 37]
v2d_3_valid_vex_list = [6, 14, 38]
v2d_3_valid_vex1_list = [6, 7, 14, 15, 38, 39]
pdpd_rate_list = [0, 1, 2, 3, 4]

for vex_num0 in v2d_0_valid_vex_list:
    for vex_num1 in v2d_0_valid_vex1_list:
        if (vex_num0 == vex_num1) or (vex_num1 == vex_num0 + 1):
            continue
        for apb_num in range (2):
            for pdpd_rate in pdpd_rate_list:
                wa.cell(row,1).value = vex_num0
                wa.cell(row,2).value = vex_num1
                wa.cell(row,3).value = 0
                wa.cell(row,4).value = apb_num
                wa.cell(row,5).value = pdpd_rate
                row+=1

for vex_num0 in v2d_1_valid_vex_list:
    for vex_num1 in v2d_1_valid_vex1_list:
        if (vex_num0 == vex_num1) or (vex_num1 == vex_num0 + 1):
            continue
        for apb_num in range (2):
            for pdpd_rate in pdpd_rate_list:
                wa.cell(row,1).value = vex_num0
                wa.cell(row,2).value = vex_num1
                wa.cell(row,3).value = 1
                wa.cell(row,4).value = apb_num
                wa.cell(row,5).value = pdpd_rate
                row+=1

for vex_num0 in v2d_2_valid_vex_list:
    for vex_num1 in v2d_2_valid_vex1_list:
        if (vex_num0 == vex_num1) or (vex_num1 == vex_num0 + 1):
            continue
        for apb_num in range (2):
            for pdpd_rate in pdpd_rate_list:
                wa.cell(row,1).value = vex_num0
                wa.cell(row,2).value = vex_num1
                wa.cell(row,3).value = 2
                wa.cell(row,4).value = apb_num
                wa.cell(row,5).value = pdpd_rate
                row+=1

for vex_num0 in v2d_3_valid_vex_list:
    for vex_num1 in v2d_3_valid_vex1_list:
        if (vex_num0 == vex_num1) or (vex_num1 == vex_num0 + 1):
            continue
        for apb_num in range (2):
            for pdpd_rate in pdpd_rate_list:
                wa.cell(row,1).value = vex_num0
                wa.cell(row,2).value = vex_num1
                wa.cell(row,3).value = 3
                wa.cell(row,4).value = apb_num
                wa.cell(row,5).value = pdpd_rate
                row+=1


sheet_name= wb['Sheet']
sheet_name.title ='underrun_C_extraterms'

wb.save("underrun_C_extraterms_list.xlsx")

