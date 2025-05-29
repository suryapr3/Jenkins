#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
wb = openpyxl.Workbook()

wa = wb.active
header = ["v2d_id" ,"apb_num", "rcs_core"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
row += 1


for v2d_num in range (4):
    for apb_num in range(2):
        for rcs_core in range (4):
            wa.cell(row,1).value = v2d_num
            wa.cell(row,2).value = apb_num
            wa.cell(row,3).value = rcs_core
            row+=1

sheet_name= wb['Sheet']
sheet_name.title ='interrupt_force_tests_list'

wb.save("interrupt_force_tests_list.xlsx")

