#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("tx_alarm_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
fail_cnt = 0

for i in range(2,row_count +1):
    v2d_num = wb1.cell(row=i,column=1).value
    apb_num = wb1.cell(row=i,column=2).value
    rcs_core = wb1.cell(row=i,column=3).value
    rv = subprocess.call(['python tx_alarm.py '  f'{v2d_num} '  f'{apb_num} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
