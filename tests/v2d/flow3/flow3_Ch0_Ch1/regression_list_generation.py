#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

V2D = fw.ABC_V2D()

wa = wb.active
header = ["tx0_vex_id" ,"rx0_vex_id" ,"tx1_vex_id" ,"rx1_vex_id" ,"v2d_id" ,"drf_clock" ,"test_config" ,"post_fliter_mode"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
wa.cell(row, 6).value = header[5]
wa.cell(row, 7).value = header[6]
wa.cell(row, 8).value = header[7]

row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

V2D.drf_clock_check(drf_clock)
print("all parameters are proper\n")

v2d_0_valid_vex_list = [0, 1, 8, 9, 32]
v2d_1_valid_vex_list = [2, 3, 10, 11, 34]
v2d_2_valid_vex_list = [4, 5, 12, 13, 36]
v2d_3_valid_vex_list = [6, 7, 14, 15, 38]
test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

for tx0_vex_num in v2d_0_valid_vex_list:
    for rx0_vex_num in v2d_0_valid_vex_list:
        if(tx0_vex_num == rx0_vex_num):
            continue
        for tx1_vex_num in v2d_0_valid_vex_list:
            for rx1_vex_num in v2d_0_valid_vex_list:
                if(tx0_vex_num == rx1_vex_num) or (tx1_vex_num == rx1_vex_num):
                    continue
                if (tx0_vex_num == tx1_vex_num) or (tx1_vex_num == rx0_vex_num) or (rx0_vex_num == rx1_vex_num) or (tx0_vex_num == rx1_vex_num):
                    continue
                for test_config in test_config_list:
                    for post_filter_mode in range(2):
                        wa.cell(row,1).value = tx0_vex_num
                        wa.cell(row,2).value = rx0_vex_num
                        wa.cell(row,3).value = tx1_vex_num
                        wa.cell(row,4).value = rx1_vex_num
                        wa.cell(row,5).value = 0
                        wa.cell(row,6).value = drf_clock
                        wa.cell(row,7).value = test_config
                        wa.cell(row,8).value = post_filter_mode
                        row+=1

for tx0_vex_num in v2d_1_valid_vex_list:
    for rx0_vex_num in v2d_1_valid_vex_list:
        if(tx0_vex_num == rx0_vex_num):
            continue
        for tx1_vex_num in v2d_1_valid_vex_list:
            for rx1_vex_num in v2d_1_valid_vex_list:
                if(tx0_vex_num == rx1_vex_num) or (tx1_vex_num == rx1_vex_num):
                    continue
                if (tx0_vex_num == tx1_vex_num) or (tx1_vex_num == rx0_vex_num) or (rx0_vex_num == rx1_vex_num) or (tx0_vex_num == rx1_vex_num):
                    continue
                for test_config in test_config_list:
                    for post_filter_mode in range(2):
                        wa.cell(row,1).value = tx0_vex_num
                        wa.cell(row,2).value = rx0_vex_num
                        wa.cell(row,3).value = tx1_vex_num
                        wa.cell(row,4).value = rx1_vex_num
                        wa.cell(row,5).value = 1
                        wa.cell(row,6).value = drf_clock
                        wa.cell(row,7).value = test_config
                        wa.cell(row,8).value = post_filter_mode
                        row+=1

for tx0_vex_num in v2d_2_valid_vex_list:
    for rx0_vex_num in v2d_2_valid_vex_list:
        if(tx0_vex_num == rx0_vex_num):
            continue
        for tx1_vex_num in v2d_2_valid_vex_list:
            for rx1_vex_num in v2d_2_valid_vex_list:
                if(tx0_vex_num == rx1_vex_num) or (tx1_vex_num == rx1_vex_num):
                    continue
                if (tx0_vex_num == tx1_vex_num) or (tx1_vex_num == rx0_vex_num) or (rx0_vex_num == rx1_vex_num) or (tx0_vex_num == rx1_vex_num):
                    continue
                for test_config in test_config_list:
                    for post_filter_mode in range(2):
                        wa.cell(row,1).value = tx0_vex_num
                        wa.cell(row,2).value = rx0_vex_num
                        wa.cell(row,3).value = tx1_vex_num
                        wa.cell(row,4).value = rx1_vex_num
                        wa.cell(row,5).value = 2
                        wa.cell(row,6).value = drf_clock
                        wa.cell(row,7).value = test_config
                        wa.cell(row,8).value = post_filter_mode
                        row+=1

for tx0_vex_num in v2d_3_valid_vex_list:
    for rx0_vex_num in v2d_3_valid_vex_list:
        if(tx0_vex_num == rx0_vex_num):
            continue
        for tx1_vex_num in v2d_3_valid_vex_list:
            for rx1_vex_num in v2d_3_valid_vex_list:
                if(tx0_vex_num == rx1_vex_num) or (tx1_vex_num == rx1_vex_num):
                    continue
                if (tx0_vex_num == tx1_vex_num) or (tx1_vex_num == rx0_vex_num) or (rx0_vex_num == rx1_vex_num) or (tx0_vex_num == rx1_vex_num):
                    continue
                for test_config in test_config_list:
                    for post_filter_mode in range(2):
                        wa.cell(row,1).value = tx0_vex_num
                        wa.cell(row,2).value = rx0_vex_num
                        wa.cell(row,3).value = tx1_vex_num
                        wa.cell(row,4).value = rx1_vex_num
                        wa.cell(row,5).value = 3
                        wa.cell(row,6).value = drf_clock
                        wa.cell(row,7).value = test_config
                        wa.cell(row,8).value = post_filter_mode
                        row+=1

sheet_name= wb['Sheet']
sheet_name.title ='flow3_Ch0_Ch1_list'

wb.save("flow3_Ch0_Ch1_list.xlsx")

