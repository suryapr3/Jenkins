#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

wa = wb.active
header = ["drf_clock" ,"test_config"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python v2d_all_regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

for test_config in test_config_list:
        wa.cell(row,1).value = drf_clock
        wa.cell(row,2).value = test_config
        row+=1

sheet_name= wb['Sheet']
sheet_name.title ='flow3_Ch0_Ch1_v2d_all_list'

wb.save("flow3_Ch0_Ch1_v2d_all_list.xlsx")

