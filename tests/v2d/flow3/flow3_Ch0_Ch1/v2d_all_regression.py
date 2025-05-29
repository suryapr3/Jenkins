#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("flow3_Ch0_Ch1_v2d_all_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0
rv = 0
fail_cnt = 0

for i in range(2,row_count +1):
    drf_clock = wb1.cell(row=i,column=1).value
    test_config = wb1.cell(row=i,column=2).value
    subprocess.call(['python3 flow3_Ch0_Ch1_v2d_all_test.py '  f'{drf_clock} ' f'{test_config} '], shell=True)

    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
