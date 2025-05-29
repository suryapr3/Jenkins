#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

V2D = fw.ABC_V2D()
wa = wb.active
header = ["quad_instance", "channel0", "channel1", "channel2", "drf_clock", "test_config"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
wa.cell(row, 6).value = header[5]
row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

if (1 == drf_clock):
    test_config_list = [0, 1, 2, 3, 5, 6, 7, 11, 12, 13]
else:
    test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

for quad_instance in range(2):
    for channel0 in range(4):
        for channel1 in range(4):
            if (channel0 == channel1):
                continue
            for channel2 in range(4):
                if (channel0 == channel2) or (channel1 == channel2):
                    continue
                for test_config in test_config_list:
                    wa.cell(row,1).value = quad_instance
                    wa.cell(row,2).value = channel0
                    wa.cell(row,3).value = channel1
                    wa.cell(row,4).value = channel2
                    wa.cell(row,5).value = drf_clock
                    wa.cell(row,6).value = test_config
                    row+=1

sheet_name= wb['Sheet']
sheet_name.title ='flow3_pdpd_ingress_repeat_3_ch'

wb.save("flow3_pdpd_ingress_repeat_3_ch_list.xlsx")

