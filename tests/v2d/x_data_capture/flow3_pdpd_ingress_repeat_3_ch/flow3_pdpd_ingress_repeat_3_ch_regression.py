#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import time
wb = openpyxl.load_workbook("flow3_pdpd_ingress_repeat_3_ch_list.xlsx")


wb1 = wb.active
row_count = wb1.max_row
print(row_count)
column_count = wb1.max_column
regression_count = 0

for i in range(2,row_count +1):
    quad_instance = wb1.cell(row=i,column=1).value
    channel0 = wb1.cell(row=i,column=2).value
    channel1 = wb1.cell(row=i,column=3).value
    channel2 = wb1.cell(row=i,column=4).value
    drf_clock = wb1.cell(row=i,column=5).value
    test_config = wb1.cell(row=i,column=6).value
    subprocess.call(['python flow3_pdpd_ingress_repeat_3_ch.py '  f'{quad_instance} ' f'{channel0} ' f'{channel1} ' f'{channel2} ' f'{drf_clock} ' f'{test_config} '], shell=True)
    regression_count = regression_count + 1
    print(f"TC completed:{regression_count}\n")
    time.sleep(100/1000)

print("regression completed\n");
