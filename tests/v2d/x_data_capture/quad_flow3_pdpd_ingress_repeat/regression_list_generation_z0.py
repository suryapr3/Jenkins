#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

V2D = fw.ABC_V2D()
wa = wb.active
header = ["quad_instance", "drf_clock", "test_config"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

V2D.drf_clock_check(drf_clock)
print("all parameters are proper\n")

if (1 == drf_clock):
    test_config_list = [0, 1, 2, 3, 5, 6, 7, 11, 12, 13]
else:
    test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

for quad_instance in range(2):
    for test_config in test_config_list:
        wa.cell(row,1).value = quad_instance
        wa.cell(row,2).value = drf_clock
        wa.cell(row,3).value = test_config
        row+=1

sheet_name= wb['Sheet']
sheet_name.title ='quad_flow3_ingress_repeat'

wb.save("quad_flow3_ingress_repeat_list.xlsx")

