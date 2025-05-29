#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

V2D = fw.ABC_V2D()
wa = wb.active
header = ["tx_vex_id" ,"rx_vex_id" ,"v2d_id" ,"apb_id" ,"drf_clock" ,"test_config" ,"post_fliter_mode", "delay_value", "capture_delay_value", "blocks_to_capture"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
wa.cell(row, 6).value = header[5]
wa.cell(row, 7).value = header[6]
wa.cell(row, 8).value = header[7]
wa.cell(row, 9).value = header[8]
wa.cell(row, 10).value = header[9]
row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

v2d_0_valid_vex_list = [0, 1, 8, 9, 32]
v2d_1_valid_vex_list = [2, 3, 10, 11, 34]
v2d_2_valid_vex_list = [4, 5, 12, 13, 36]
v2d_3_valid_vex_list = [6, 7, 14, 15, 38]
if (0 == drf_clock):
    capture_delay_value_list = [0, 512, 1024]
    blocks_to_compare_list = [50, 70]
else:
    capture_delay_value_list = [0, 768, 1536]
    blocks_to_compare_list = [30, 40]
test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
delay_value_list = [768, 1152]
post_filter_mode = 0

for tx_vex_num in v2d_0_valid_vex_list:
    for rx_vex_num in v2d_0_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for test_config in test_config_list:
                for capture_delay_value in capture_delay_value_list:
                    for blocks_to_compare in blocks_to_compare_list:
                        wa.cell(row,1).value = tx_vex_num
                        wa.cell(row,2).value = rx_vex_num
                        wa.cell(row,3).value = 0
                        wa.cell(row,4).value = apb_num
                        wa.cell(row,5).value = drf_clock
                        wa.cell(row,6).value = test_config
                        wa.cell(row,7).value = post_filter_mode
                        wa.cell(row,8).value = delay_value_list[drf_clock]
                        wa.cell(row,9).value = capture_delay_value
                        wa.cell(row,10).value = blocks_to_compare
                        row+=1


for tx_vex_num in v2d_1_valid_vex_list:
    for rx_vex_num in v2d_1_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for test_config in test_config_list:
                for capture_delay_value in capture_delay_value_list:
                    for blocks_to_compare in blocks_to_compare_list:
                        wa.cell(row,1).value = tx_vex_num
                        wa.cell(row,2).value = rx_vex_num
                        wa.cell(row,3).value = 1
                        wa.cell(row,4).value = apb_num
                        wa.cell(row,5).value = drf_clock
                        wa.cell(row,6).value = test_config
                        wa.cell(row,7).value = post_filter_mode
                        wa.cell(row,8).value = delay_value_list[drf_clock]
                        wa.cell(row,9).value = capture_delay_value
                        wa.cell(row,10).value = blocks_to_compare
                        row+=1

for tx_vex_num in v2d_2_valid_vex_list:
    for rx_vex_num in v2d_2_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for test_config in test_config_list:
                for capture_delay_value in capture_delay_value_list:
                    for blocks_to_compare in blocks_to_compare_list:
                        wa.cell(row,1).value = tx_vex_num
                        wa.cell(row,2).value = rx_vex_num
                        wa.cell(row,3).value = 2
                        wa.cell(row,4).value = apb_num
                        wa.cell(row,5).value = drf_clock
                        wa.cell(row,6).value = test_config
                        wa.cell(row,7).value = post_filter_mode
                        wa.cell(row,8).value = delay_value_list[drf_clock]
                        wa.cell(row,9).value = capture_delay_value
                        wa.cell(row,10).value = blocks_to_compare
                        row+=1


for tx_vex_num in v2d_3_valid_vex_list:
    for rx_vex_num in v2d_3_valid_vex_list:
        if(tx_vex_num == rx_vex_num):
            continue
        for apb_num in range (2):
            for test_config in test_config_list:
                for capture_delay_value in capture_delay_value_list:
                    for blocks_to_compare in blocks_to_compare_list:
                        wa.cell(row,1).value = tx_vex_num
                        wa.cell(row,2).value = rx_vex_num
                        wa.cell(row,3).value = 3
                        wa.cell(row,4).value = apb_num
                        wa.cell(row,5).value = drf_clock
                        wa.cell(row,6).value = test_config
                        wa.cell(row,7).value = post_filter_mode
                        wa.cell(row,8).value = delay_value_list[drf_clock]
                        wa.cell(row,9).value = capture_delay_value
                        wa.cell(row,10).value = blocks_to_compare
                        row+=1

sheet_name= wb['Sheet']
sheet_name.title ='flow3_pdpd_ingress_duration'

wb.save("flow3_pdpd_ingress_duration_list.xlsx")

