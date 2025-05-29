#!/usr/intel/bin/python3.6.3a
import UsrIntel.R3
import openpyxl
import subprocess
import sys
import scripts as fw
wb = openpyxl.Workbook()

V2D = fw.ABC_V2D()
wa = wb.active
header = ["quad_instance", "drf_clock", "test_config", "delay_value", "capture_delay_value", "delay_antenna_capture"]
row = 1

wa.cell(row, 1).value = header[0]
wa.cell(row, 2).value = header[1]
wa.cell(row, 3).value = header[2]
wa.cell(row, 4).value = header[3]
wa.cell(row, 5).value = header[4]
wa.cell(row, 6).value = header[5]
row += 1

arg_len = len(sys.argv)
if arg_len != 2:
    print("Script should be run in below format\n")
    print("python regression_list_generation.py <drf_clock>\n")
    print("As there were no inputs provided executing with drf_clock: 0\n")
    drf_clock = 0
else:
    drf_clock = (int) (sys.argv[1])

V2D.drf_clock_check(drf_clock)
print("all parameters are proper\n")

test_config_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

if (0 == drf_clock):
    capture_delay_value_list = [0, 512, 1024]
    delay_antenna_capture_list = [0, 512, 1024]
else:
    capture_delay_value_list = [0, 768, 1536]
    delay_antenna_capture_list = [0, 768, 1536]

delay_value_list = [768, 1152]

for quad_instance in range(2):
    for test_config in test_config_list:
        for capture_delay_value in capture_delay_value_list:
            for delay_antenna_capture_value in delay_antenna_capture_list:
                wa.cell(row,1).value = quad_instance
                wa.cell(row,2).value = drf_clock
                wa.cell(row,3).value = test_config
                wa.cell(row,4).value = delay_value_list[drf_clock]  #delay value
                wa.cell(row,5).value = capture_delay_value
                wa.cell(row,6).value = delay_antenna_capture_value
                row+=1

sheet_name= wb['Sheet']
sheet_name.title ='quad_flow3_pdpd_ingress'

wb.save("quad_flow3_pdpd_ingress_list.xlsx")

