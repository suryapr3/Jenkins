# INTEL CONFIDENTIAL
#
# Copyright 2022 Intel Corporation All Rights Reserved.
#
# The source code contained or described herein and all documents related
# to the source code ("Material") are owned by Intel Corporation or its
# suppliers or licensors. Title to the Material remains with Intel
# Corporation or its suppliers and licensors. The Material contains trade
# secrets and proprietary and confidential information of Intel or its
# suppliers and licensors. The Material is protected by worldwide copyright
# and trade secret laws and treaty provisions. No part of the Material
# may be used, copied, reproduced, modified, published, uploaded, posted,
# transmitted, distributed, or disclosed in any way without Intel's prior
# express written permission.
#
# No license under any patent, copyright, trade secret or other
# intellectual property right is granted to or conferred upon you by
# disclosure or delivery of the Materials, either expressly, by
# implication, inducement, estoppel or otherwise. Any license under such
# intellectual property rights must be express and approved by Intel in
# writing.

import sys
from pathlib import Path # for file name/path

# for working with speadsheet:
from openpyxl import Workbook
from openpyxl import load_workbook

# for working with csv:
import csv

import tempfile

from .util import *
from .table_info import *
from .style import tbl_cell_style
from .build_headers import *
from .build_html import build_html_output
from .build_xml import build_xml_output
from .build_table_init import build_table_init
from .build_test import build_table_test
from pm.access_macros import build_access_macros_output

def x2c_convert(xlsx_file, worksheets, dest_dir, do_msg):
    """xlsx to csv convertion

    Converting .xlsx file into .csv file

    Parameters
    ----------
    xlsx_file : str
        the name of xlsx file
    worksheets : set
        a set of Worksheet, the name of each sheet to be converted. the specified sheets must exist in the workbook
    dest_dir : str
        the name of path to store resulting .csv file
    do_msg : str
        name for message header, when not None, indicating doing message table process

    Returns
    -------
    None
    """
    if Path(xlsx_file).is_file() and xlsx_file.endswith(".xlsx"):
        table_pos_info = get_table_pos_info()
        wb = load_workbook(xlsx_file, data_only = True)
        if worksheets == None:
            valid_sheets = wb.sheetnames
        else:
            valid_sheets = set(worksheets) & set(wb.sheetnames)
            if not valid_sheets:
                p_red2('No worksheet provided in command line found in file', xlsx_file)
                sys.exit()
            if len(worksheets) > len(valid_sheets):
                p_red2('Not all sheets provided in command line found in file', xlsx_file)
        if 'address_size_layout' in wb.sheetnames:
            # only generate .csv from tables specified in address_size_layout
            tbl_name_set = set(get_table_list(wb['address_size_layout']))
            tbl_name_set |= {'address_size_layout'}
            if do_msg != None:
                p_red2('For message table: no need of address_size_layout in file', xlsx_file)
        else:
            p_purple_red6('Critical sheet', "address_size_layout", "is missing from file", xlsx_file, ': for message header generation (see --msg option) this is', 'OK')
            tbl_name_set = set()
            for sn in valid_sheets:
                tbl_name_set.add(sn.lower())

        path = Path(dest_dir)
        path.mkdir(parents=True, exist_ok=True)

        # deleting any existing .csv files first in destination folder:
        for f in path.glob('*.csv'):
            try:
                f.unlink()
            except OSError as e:
                p_red4("Error", e.strerror, "happened when deleting file", f)

        for sheet in valid_sheets:
            ws = wb[sheet]
            if ws.title == 'address_size_layout':
                fn_name = 'address_size_layout'
            else:
                fn_name = ws[table_pos_info[0]].value # use table_name as csv file name
            if fn_name != None and fn_name.lower() in tbl_name_set:
                csv_fn = fn_name.lower()
                csvfile= dest_dir + '/' + csv_fn + '.csv'
                with open(csvfile, "w", newline="") as out_f:
                    col = csv.writer(out_f, delimiter = ",", lineterminator="\n")
                    for row in ws.iter_rows():
                        col.writerow([cell.value for cell in row])
                p_green4("Generated", csvfile, 'from', xlsx_file)
            else:
                p_red2("skipping not used sheet", wb[sheet].title)
    else:
        p_red2("Skipping ...either file not exist, or wrong type. Please check given file:", xlsx_file)

def c2x_convert(csv_input, xlsx_file_name, dest_dir):
    """.csv to xlsx convertion

    Convertig .csv file to .xlsx file

    Parameters
    ----------
    csv_input : str
        folder name containning .csv file
    xlsx_file_name : str
        specifying the name of result .xlsx file to be saved
    dest_dir : str
        the name of path to store resulting .xlsx file

    Returns
    -------
    None
    """
    xlsx_file_name = xlsx_file_name + ('.xlsx', '')[xlsx_file_name.endswith(".xlsx")]
    xlsx_file = dest_dir + '/' + xlsx_file_name
    if Path(xlsx_file).is_file(): # xlsx file exists
        wb = load_workbook(xlsx_file, data_only=True) # Load our Excel File
    else: # if xlsx file doesn't exist
        wb = Workbook() # create a new workbook
        wb.active.title = 'blank_workbook_first_sheet' # indicating the first sheet from an empty workbook
    var_csv_files = []
    csv_list = []
    for csv_path_dir in csv_input:
        csv_path = Path(csv_path_dir)
        if csv_path.is_dir():
            files_in_dir = csv_path.iterdir()
            if csv_list != []:
                csv_files = set([csv_path_dir + '/' + f.name for f in files_in_dir \
                        if f.is_file() and f.name.endswith(".csv") and f.name not in var_csv_files and f.name in csv_list])
            else:
                csv_files = set([csv_path_dir + '/' + f.name for f in files_in_dir \
                        if f.is_file() and f.name.endswith(".csv") and f.name not in var_csv_files])
        csv_files_1 = set([csv_path_dir + '/' + f for f in var_csv_files])
        num = 0 # to track how many .csv files really processed
        for f in set(csv_files) | csv_files_1: # removing duplicates just in case
            #print(Path(f).resolve().parent)
            #print(f)
            if Path(f).is_file() and f.endswith(".csv"):
                name = Path(f).stem
                if name in wb.sheetnames: # replacing so keep same idx
                    sheet = wb[name] # get sheet by name
                    old_idx = wb.index(sheet) # get sheet index
                    wb.remove(sheet)
                    ws = wb.create_sheet(name, old_idx)
                    #ws_from_idx = wb.worksheets[idx] # get worksheet by index
                else:
                    if wb.active.title == 'blank_workbook_first_sheet':
                        wb.active.title = name
                        ws = wb.active
                    else:
                        ws = wb.create_sheet(name)
                with open(f, newline="", encoding="utf8") as csv_in:
                    for row in csv.reader(csv_in, delimiter=","):
                        ws.append([cell2num(cell) for cell in row])
                if ws.title != 'address_size_layout':
                    #print(ws.title)
                    tbl_cell_style(ws)
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToHeight = False
                num += 1
        if num > 0:
            path = Path(dest_dir)
            path.mkdir(parents=True, exist_ok=True)
            wb.save(xlsx_file) #(csv_files.replace(".csv", ".xlsx"))
            p_green4("Generated ", xlsx_file, 'from', str(num) + ' .csv files')
        else:
            p_red("No .csv files processed, please check given .csv file path!!!")
            # no .csv file processed

def cell2num(cell):
    """convert csv cell into number

    Converting .csv cell into int or float based on input cell is number or not

    Parameters
    ----------
    cell : str
        a specific cell from .csv file

    Returns
    -------
    cell : int if input cell is numertic type
         : float if input cell is NOT numertic type
    """
    if cell.isnumeric():
        return int(cell)
    try:
        return float(cell)
    except ValueError:
        pass
    return cell

def x2h_convert(xlsx_file, sheets, dest_dir, no_test, use_xtensa, addr_sec, length, ofs, xml, html_dir, table_init, test_file, tb_access_macros, do_msg, hdr_list, ctypes_dir, no_msg_dic_gen):
    """convert xlsx into C header file

    extracting table info from xlsx file, then generating C structure definitions and output as C header files

    Parameters
    ----------
    xlsx_file : str
        input xlsx file name (including path)
	sheets : set
        a set of Worksheet to be processed, if not specified, all sheets inside the workbook, if listed in sheet 'address_size_layout'  will be processed
	dest_dir : str
        the destination folder to hold the output files
	no_test : bool
        true to generate test C file, false no test file being generated
	use_xtensa : bool
        toolset selection: true to use XTENSA toolset, false to use GCC toolset.
    addr_sec : str
        start address or memory section name to put the combined table
    length : str
        string contains memory size to hold the combined table
    ofs : str
        string contains the offset to the start address
	xml : str
        specify the .xsd file to be used with generated XML file
	html_dir : str
        specifying the folder to hold the generated HTML files
    table_init : str
        name of table init file
    test_file : str
        name of table test file
    tb_access_macros : str
        header file name of table access macros definition
    do_msg : str
        name for message header, when not None, indicating doing message table process
    hdr_list : list
        list of user specified header files to be included
    ctypes_dir: str
        folder to store python ctypes compatible table struct definitions, an empty string means no generation of these definitions

    Returns
    ----------
    None
    """
    if xlsx_file.endswith(".xlsx"):
        p_green2("Generating header files from", xlsx_file)
        wb = load_workbook(xlsx_file, data_only=True) # Load our Excel File
        pos_shift = 36 # offset of ':' in bit field layout in struct, this implies bit field name length as 36
        if not 'address_size_layout' in wb.sheetnames:
            if do_msg == None:
                p_purple_red6('Critical sheet', "address_size_layout", "not found in", xlsx_file, '', '')
                sys.exit()
            else:
                table_dim_dic = {}
                super_table_name = do_msg
        else:
            (super_table_name, addr_len_tup, table_dim_dic) = build_layout_info(wb['address_size_layout'])
            if do_msg != None: # this should not happen, but just to enforce
                super_table_name = do_msg
        valid_sheets = set()
        if sheets == None:
            #valid_sheets = set(wb.sheetnames) & set(table_dim_dic.keys())
            for sn in wb.sheetnames:
                if do_msg == None:
                    if sn.lower() in set(table_dim_dic.keys()): # XXX forcing lower case for tbl_name, check build_layout_info
                        valid_sheets.add(sn)
                else:
                    valid_sheets.add(sn.lower())
        else:
            #valid_sheets = set(sheets) & set(wb.sheetnames) & set(table_dim_dic.keys())
            for sn in set(sheets) & set(wb.sheetnames):
                if do_msg == None:
                    if sn.lower() in set(table_dim_dic.keys()): # XXX forcing lower case for tbl_name, check build_layout_info
                        valid_sheets.add(sn)
                else:
                    valid_sheets.add(sn.lower())
            if not valid_sheets:
                p_red("No worksheet provided in commandline found in workbook or in address_size_layout, exiting ...")
                sys.exit()
            if len(valid_sheets) < len(sheets):
                p_red("Not all sheets provided in commandline found in workbook...\n")
        if valid_sheets == 'address_size_layout':
            sys.exit()
        else:
            valid_sheets.discard('address_size_layout') # address_size_layout already processed

        # info from sheet address_size_layout:
        # super_table_name
        # addr_len_tup
        # table_dim_dic = {tbl_name:(A, B), ...}
        #
        # info from table sheet:
        # tbl_dic[tbl_name] = (tbl_des, bf_blks, sub_tbl_dic, tbl_bitwidth_tup, align, category, subcategory)
        # bf_blks = [[bf_blk], [], ...]
        #   bf_blk = [[bf_info], [], ...]
        #       bf_info = ('name', 'description', 'unit', 'granularity', 'min', 'max', 'sign', 'bit width', 'upper position')
        # sub_tbl_dic = {sub_tbl_name:(L, R, sub_tbl_hex, sub_tbl_bf_info), ...}
        # sub_tbl_hex = (hex, ...)
        # sub_tbl_bf_info = ((field0_val, field1_val, ...), ...)
        # tbl_bitwidth_tup = (tbl_bitwidth, type_bitwidth, tbl_struct_blk_num, bf_max_len)
        # enum_dic = {enum_type_name:[enum_elem, ...], }

        xlsx_updt, tbl_dic, enum_dic = process_workbook(wb, valid_sheets, table_dim_dic)
        if xlsx_updt:
            wb.save(xlsx_file)
        table_sanity_check(tbl_dic)

        path = Path(dest_dir)
        path.mkdir(parents=True, exist_ok=True)
        if ctypes_dir:
            path = Path(ctypes_dir)
            path.mkdir(parents=True, exist_ok=True)
        if enum_dic != {}:
            ext = '_enum'
            with open(dest_dir + '/' + super_table_name.lower() + ext + '.h', "w") as out_f:
                build_copyright_info(out_f)
                build_includ_head_info(out_f, super_table_name, ext)
                build_header_enum(out_f, enum_dic, super_table_name, ctypes_dir)
                build_includ_tail_info(out_f, super_table_name, ext)
                p_green2("Generated header file", super_table_name + ext + '.h')

        if do_msg == None:
            ext = '_bitfield'
        else:
            ext = '_msg'
        with open(dest_dir + '/' + super_table_name.lower() + ext + '.h', "w") as out_f:
            build_copyright_info(out_f)
            build_includ_head_info(out_f, super_table_name, ext)
            build_header_bitfield(out_f,super_table_name, tbl_dic, enum_dic, do_msg, hdr_list, ctypes_dir, no_msg_dic_gen)
            build_includ_tail_info(out_f, super_table_name, ext)
            p_green2("Generated header file", super_table_name + ext + '.h')

        if html_dir != None:
            build_html_output(super_table_name, tbl_dic, table_dim_dic, html_dir, do_msg)

        if do_msg == None:
            ext = '_const'
            file_name =Path(xlsx_file).stem
            with open(dest_dir + '/' + super_table_name.lower() + ext + '.h', "w") as out_f:
                build_copyright_info(out_f)
                build_includ_head_info(out_f, super_table_name.lower(), ext)
                build_table_consts(out_f, super_table_name, table_dim_dic, pos_shift + 16)
                build_includ_tail_info(out_f, super_table_name, ext)
            p_green2("Generated header file", super_table_name + ext + '.h')

            ext = '_data'
            with open(dest_dir + '/' + super_table_name.lower() + ext + '.h', "w") as out_f:
                build_copyright_info(out_f)
                build_includ_head_info(out_f, super_table_name, ext)
                build_header_data(out_f, super_table_name, table_dim_dic, tbl_dic, pos_shift, ext)
                build_includ_tail_info(out_f, super_table_name, ext)
                p_green2("Generated header file", super_table_name + ext + '.h')

            if tb_access_macros != None:
                build_access_macros_output(tb_access_macros, dest_dir)

            ext = '_table'
            with open(dest_dir + '/' + super_table_name.lower() + ext + '.h', "w") as out_f:
                build_copyright_info(out_f)
                build_includ_head_info(out_f, super_table_name, ext)
                build_header_super_table(out_f, super_table_name, tbl_dic, table_dim_dic, pos_shift, tb_access_macros, addr_len_tup, addr_sec, ofs, ctypes_dir)
                build_includ_tail_info(out_f, super_table_name, ext)
                p_green2("Generated header file", super_table_name + ext + '.h')

            if xml != None:
                build_xml_output(super_table_name, super_table_name, tbl_dic, table_dim_dic, addr_len_tup, xml)

            if table_init != None:
                with open(table_init, "w") as out_f:
                    build_copyright_info(out_f)
                    build_table_init(out_f, tbl_dic, super_table_name.lower(), addr_sec, pos_shift)
                    p_green2("Generated table init file", table_init)

            if no_test == False:
                if test_file == None:
                    tmp_dir = tempfile.mktemp()
                    test_file = tmp_dir + '_' + super_table_name.lower()
                if test_file.endswith(".c"):
                    test_file = test_file[:-2]
                build_table_test(test_file + '.c', tbl_dic, super_table_name, table_dim_dic, addr_len_tup, addr_sec, length, ofs, use_xtensa, pos_shift, table_init, tb_access_macros)
                p_purple3("\nPlease refer to test file", test_file + '.c',  "on how to initialize/access table\n")
                if not use_xtensa:
                    import subprocess
                    try:
                        subprocess.check_call("{0} {1} {2} {1}.c {3}".format('gcc -std=c99 -Wall -o', test_file, " -I" + dest_dir + " -Ifw/com/inc", (table_init, '')[table_init == None]), shell = True)
                    except subprocess.CalledProcessError:
                        p_red2('Failed to compile header file', test_file + '.c\n')
                        sys.exit()
                    else:
                        p_green2("Running test program:", test_file)
                        run_ret = subprocess.run([test_file])
                        yesno = input("delete both " + test_file + ' and ' + test_file + '.c, Y|n?').lower()
                        yes={'', 'y', 'ye', 'yes'}
                        if yesno in yes:
                            p_red2("deleting test source file:", test_file + '.c')
                            Path(test_file + '.c').unlink()
                            p_red2("deleting test executable:", test_file)
                            Path(test_file).unlink()
    else:
        p_red2("Skipping ... expect .xlsx file but given", xlsx_file)


