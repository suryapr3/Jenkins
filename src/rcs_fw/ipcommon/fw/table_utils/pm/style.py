# INTEL CONFIDENTIAL
#
# Copyright 2022 Intel Corporation All Rights Reserved.
#
# The source code contained or described herein and all documents related
# to the source code ("Material") are owned by Intel Corporation or its
# suppliers or licensors. Title to the Material remains with Intel
# Corporation or its suppliers and licensors. The Material contains trade
# secrets and proprietary and confidential information of Intel or its
# suppliers and licensors. The Material is protected by worldwide copyright
# and trade secret laws and treaty provisions. No part of the Material
# may be used, copied, reproduced, modified, published, uploaded, posted,
# transmitted, distributed, or disclosed in any way without Intel's prior
# express written permission.
#
# No license under any patent, copyright, trade secret or other
# intellectual property right is granted to or conferred upon you by
# disclosure or delivery of the Materials, either expressly, by
# implication, inducement, estoppel or otherwise. Any license under such
# intellectual property rights must be express and approved by Intel in
# writing.

import sys
import re # for regular expression
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import BLUE, BLACK #, GREEN, YELLOW, DARKBLUE, DARKGREEN
from openpyxl.utils.cell import get_column_letter

from .table_info import get_table_pos_info
from .util import parse_dim_info, p_red4

def tbl_cell_style(ws):
    """applying customized styles to cell

    applying table cell style (eg, filling color or line stype) to different table elements for easy reviewing

    Parameters
    ----------
    ws : Worksheet
        worksheet to apply the cell style

    Returns
    -------
    None
    """
    fill_red = [PatternFill(patternType='solid', fgColor='EC2C33'), PatternFill(patternType='solid', fgColor='EC2C33')]
    fill_blue = [PatternFill(patternType='solid', fgColor='03FCF4'), PatternFill(patternType='solid', fgColor='03FCF4')]
    fill_green = [PatternFill(patternType='solid', fgColor='35FC03'), PatternFill(patternType='solid', fgColor='35FC03')]
    fill_yellow = [PatternFill(patternType='solid', fgColor='FCBA03'), PatternFill(patternType='solid', fgColor='FCBA03')]
    fill_color = [fill_green, fill_blue, fill_yellow, fill_red]

    dotted = Side(border_style="dotted", color=BLUE)
    hair = Side(border_style="hair", color=BLACK)
    thin = Side(border_style="thin", color=BLACK)
    double = Side(border_style="double", color=BLACK)

    table_pos_info = get_table_pos_info()
    tbl_bitwidth_pos = table_pos_info[2]
    tbl_struct_pos = table_pos_info[3]
    sub_table_dim_col = table_pos_info[5]
    bf_col_min = ws[tbl_struct_pos].column

    type_bitwidth = ws.cell(column=bf_col_min, row = ws[tbl_struct_pos].row - 1).value + 1 # basic data type size in bits
    if not isinstance(type_bitwidth, int):
        p_red4("missing/incorrect type bitwidth info at cell", "{}{}".format(get_column_letter(bf_col_min), ws[tbl_struct_pos].row - 1), "in sheet", ws.title)
        return

    dim = ws[tbl_bitwidth_pos].value
    if isinstance(dim, str):
        dim_tup = tuple(map(int,  re.findall(r'\d+', dim)))
        if len(dim_tup) != 2:
            p_red4("missing/incorrect table bitwidth info at cell", tbl_bitwidth_pos, "in sheet", ws.title)
            #sys.exit()
            return
        else:
            tbl_bitwidth = dim_tup[0] * dim_tup[1]
    elif isinstance(dim, int):
        tbl_bitwidth = dim
    else:
        p_red4("missing/incorrect table bitwidth info at cell", tbl_bitwidth_pos, "in sheet", ws.title)
        #sys.exit()
        return
    tbl_struct_blk_num = tbl_bitwidth // type_bitwidth
    # bit-field:
    bf_row_min = ws[tbl_struct_pos].row
    bf_info_depth = table_pos_info[4]
    for b in range(type_bitwidth):
        ws.cell(row = bf_row_min - 1, column = bf_col_min + b).alignment = Alignment(horizontal='left', vertical='top')
    bitwidth_info = []
    for blk_num in range(tbl_struct_blk_num):
        itr_col = bf_col_min
        itr_row = bf_row_min + blk_num * bf_info_depth
        i = blk_num
        bitwidth_info_blk = []
        while itr_col < type_bitwidth + bf_col_min:
            bitwidth = ws.cell(row = itr_row + 7, column = itr_col).value
            if not isinstance(bitwidth, int):
                p_red4("missing/incorrect type bitwidth info at cell", "{}{}".format(get_column_letter(itr_col), itr_row + 7), "in sheet", ws.title)
                sys.exit()
            bitwidth_info_blk.append(bitwidth)
            for rows in ws.iter_rows(min_row=itr_row, max_row=itr_row + bf_info_depth - 1, min_col=itr_col, max_col=itr_col + bitwidth - 1):
                for cell in rows:
                    cell.fill = fill_color[i % 4][blk_num % 2]
                    # PatternFill(start_color=fill_color[i%2], end_color=fill_color[i%2], fill_type = "solid") #"solid")
                    cell.alignment = Alignment(horizontal = 'left', vertical = 'top')
                    if cell.column == itr_col + bitwidth - 1:
                        cell.border = Border(top=hair, left=hair, right=double, bottom=hair)
                        if cell.row == itr_row + bf_info_depth - 1:
                            cell.border = Border(top=hair, left=hair, right=double, bottom=dotted)
                        if cell.row == bf_row_min + bf_info_depth * tbl_struct_blk_num - 1:
                            cell.border = Border(top=hair, left=hair, right=double, bottom=thin)
                    else:
                        cell.border = Border(top=hair, left=hair, right=hair, bottom=hair)
                        if cell.row == itr_row + bf_info_depth - 1:
                            cell.border = Border(top=hair, left=hair, right=hair, bottom=dotted)
                        if cell.row == bf_row_min + bf_info_depth * tbl_struct_blk_num  - 1:
                            cell.border = Border(top=hair, left=hair, right=hair, bottom=thin)
            itr_col += bitwidth
            i += 1
        bitwidth_info.append(bitwidth_info_blk)
    # data
    current_sub_tbl_row = ws[tbl_struct_pos].row + tbl_struct_blk_num * bf_info_depth
    sub_tbl_dim = ws.cell(row = current_sub_tbl_row, column = sub_table_dim_col).value
    # Note: when reading from csv, the empty cell has value as '' instead of None:
    if sub_tbl_dim == None or sub_tbl_dim == '' or not isinstance(sub_tbl_dim, str):
        return
    while sub_tbl_dim != None and sub_tbl_dim != '' and isinstance(sub_tbl_dim, str): # continue if not empty cell for dim cell
        L = max([parse_dim_info(d, 2, ws, "{}{}".format(get_column_letter(sub_table_dim_col), current_sub_tbl_row))[0]\
                for d in sub_tbl_dim.split(',')])
        for l in range(L):
            for blk_num in range(tbl_struct_blk_num):
                itr_col = bf_col_min
                itr_row = current_sub_tbl_row + blk_num + tbl_struct_blk_num * l + 1
                bitwidth_cnt = 0
                bitwidt_accum = 0
                while itr_col < type_bitwidth + bf_col_min:
                    bitwidth = bitwidth_info[blk_num][bitwidth_cnt]
                    cell = ws.cell(row = itr_row, column = itr_col)
                    cell.fill = fill_color[(bitwidth_cnt + blk_num) % 4][blk_num % 2]
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                    if cell.column == bf_col_min + bitwidth + bitwidt_accum - 1:
                        bitwidt_accum += bitwidth
                        bitwidth_cnt += 1
                        cell.border = Border(top=hair, left=hair, right=double, bottom=dotted)
                        if blk_num == tbl_struct_blk_num - 1:
                            cell.border = Border(top=hair, left=hair, right=double, bottom=thin)
                    else:
                        cell.border = Border(top=hair, left=hair, right=hair, bottom=dotted)
                        if blk_num == tbl_struct_blk_num - 1:
                            cell.border = Border(top=hair, left=hair, right=hair, bottom=thin)
                    itr_col += 1
        current_sub_tbl_row += L * tbl_struct_blk_num + 1
        sub_tbl_dim = ws.cell(row = current_sub_tbl_row, column = sub_table_dim_col).value


