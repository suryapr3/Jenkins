# INTEL CONFIDENTIAL
#
# Copyright 2022 Intel Corporation All Rights Reserved.
#
# The source code contained or described herein and all documents related
# to the source code ("Material") are owned by Intel Corporation or its
# suppliers or licensors. Title to the Material remains with Intel
# Corporation or its suppliers and licensors. The Material contains trade
# secrets and proprietary and confidential information of Intel or its
# suppliers and licensors. The Material is protected by worldwide copyright
# and trade secret laws and treaty provisions. No part of the Material
# may be used, copied, reproduced, modified, published, uploaded, posted,
# transmitted, distributed, or disclosed in any way without Intel's prior
# express written permission.
#
# No license under any patent, copyright, trade secret or other
# intellectual property right is granted to or conferred upon you by
# disclosure or delivery of the Materials, either expressly, by
# implication, inducement, estoppel or otherwise. Any license under such
# intellectual property rights must be express and approved by Intel in
# writing.

import sys
from openpyxl.utils.cell import coordinate_from_string, get_column_letter
#from openpyxl.utils.cell import get_column_letter
from .util import *

def get_table_list(address_size_layout_sheet):
    """Extract table info.

    Extract table layout, size and address info from a specific sheet.

    Parameters
    ----------
    address_size_layout_sheet : Worksheet
        Worksheet that contains table layout, size and address info

    Returns
    -------
    tbl_name_list : list
        list of table names, ordered as in layout
    """
    info = []
    for row in address_size_layout_sheet.values:
        info.append(row)
    tbl_name_list = [info[r][1] for r in range(2, len(info))]
    return tbl_name_list

def get_table_pos_info():
    """Get position info of required cells

    Get the position info of mandotery cells what are in pre-defined/fixed location, and some constants related to table position calculation

    Parameters
    ----------
    None

    Returns
    -------
    table_pos_info : set
        a set containning position info and constants related to position calculation
    """
    table_pos_info = ()
    table_pos_info += ('B1',) # tbl_name_pos = 'B1' # cell holding table name
    table_pos_info += ('B2',) # tbl_des_pos = 'B2' # cell holding table description
    table_pos_info += ('B3',) # tbl_dim_pos = 'B3' # cell holding table bitwidth info
    table_pos_info += ('D5',) # tbl_struct_pos = 'D5' # cell holding table struct bit-field definition
    table_pos_info += (8,)    # bf_info_depth = 8 # num of info entry for each bit-field
    table_pos_info += (2,)    # sub_table_dim_col = 2
    table_pos_info += ('C3',) # tbl_alignment_pos = 'C3' # cell holding table alignment info
    return table_pos_info

def build_layout_info(ws):
    """extract table address, size and layout info

    Parameters
    ----------
    ws : Worksheet
        the worksheet which contains table info for special table 'address_size_layout'

    Returns
    ----------
    tup(super_table_name, addr_len_tup, table_dim_dic)
        super_table_name: name of the grouped tables
        addr_len_tup: a tup contains start address of the table and overall length of the table
        table_dim_dic: dictionary contains table dimension info
    """
    p_green2("porcessing sheet", ws.title)
    info = []
    addr_dic = {}
    for row in ws.values:
        info.append(row)
    super_table_name = info[0][1]
    if super_table_name == None:
        p_red4("expected super table name at cell", "{}{}".format(get_column_letter(1 + 1), 1), "in sheet", ws.title )
        sys.exit()
    super_table_name = super_table_name.lower() # XXX forcing super table name to low case

    if None == info[1][1]:
        p_red2("address/length info is missing in sheet", ws.title)
        sys.exit()
    else:
        address, length = info[1][1].split(',')
    addr_len_tup = (address.strip(), length.strip())

    table_dim_dic = {}
    for r in range(2, len(info)):
        tmp_tbl_name = info[r][1]
        if tmp_tbl_name == None:
            p_red2("table name is missing at row", r + 1)
            sys.exit()
        tmp_tbl_name = tmp_tbl_name.lower() # XXX forcing table name to low case
        table_dim_dic[tmp_tbl_name] = parse_dim_info(info[r][2], 2, ws, "{}{}".format(get_column_letter(2 + 1), r + 1))
    #pp.pprint(table_dim_dic) # {tbl_name:(A, B), ...}
    return (super_table_name, addr_len_tup, table_dim_dic)

def keyinfo_check(sheet, tbl_name_pos, tbl_bitwidth_pos, tbl_struct_pos):
    """check existance of required info

    Parameters
    ----------
    sheet : Worksheet
        the worksheet which contains table info
    tbl_name_pos : str
        cell position string for table name
    tbl_bitwidth_pos : str
        cell position info for table bitwidth
    tbl_struct_pos : str
        cell position info for table struct

    Returns
    ----------
    None
    """
    (c, r) = coordinate_from_string(tbl_struct_pos)
    if sheet[tbl_name_pos].value == None:
        p_red4('Missing table name at cell', tbl_name_pos, 'in sheet', sheet.title)
        sys.exit()
    elif sheet[tbl_bitwidth_pos].value == None:
        p_red4('Missing table bithwidth info at cell', tbl_bitwidth_pos, 'in sheet', sheet.title)
    elif sheet[c + str(r - 1)].value == None:
        p_red4('Missing base type size info at cell', c + str(r - 1), 'in sheet', sheet.title)
        sys.exit()
    else:
        pass

def table_sanity_check(tbl_dic):
    """table content sanity check

    make sure no duplicated subtable

    Parameters
    ----------
    tbl_dic : dict
        a dictionary: tbl_name as key, tup (tbl_des, bf_blks, sub_tbl_dic, tbl_bitwidth_tup, align) as value
        tbl_des: table description
        bf_blks: bit-field block list, each bit-field block is a list of info for a specific bit-field
        sub_tbl_dic: a dictionary with info about a subtable
        tbl_bitwidth_tup: a tup with bitwidth info for each table
        align: specifying the alignment attribute of this table

    Returns
    ----------
    None
    """
    tbl_names_dic = {}
    for tbl_name in tbl_dic.keys():
        if tbl_dic[tbl_name][2] != {}: # do_msg != None
            for sub_tbl_name in tbl_dic[tbl_name][2].keys():
                if sub_tbl_name not in tbl_names_dic.keys():
                    tbl_names_dic[sub_tbl_name] = sub_tbl_name
                else:
                    p_purple_red6('Duplicated sub table name', sub_tbl_name, 'in table', tbl_names_dic[sub_tbl_name], 'and table', tbl_name)
                    sys.exit()

def build_bf_struct(sheet, tbl_bitwidth_pos, tbl_struct_pos, bf_info_depth, enum_dic):
    """Build up bit-fields info for each table

    Constructing C struct bit-field info for each table

    Parameters
    ----------
    sheet : Worksheet
        the worksheet which contains all bitfield info
    tbl_bitwidth_pos : str
        cell position info for table bitwidth
    tbl_struct_pos : str
        cell position info for table struct
    bf_info_depth : int
        number of info entries for each bit-field, this is a constant
    enum_dic : dict
        a dictionary: enum type name as key, a list of enum elements for this enum type as value

    Returns
    ----------
    a tupple contains below elements:
        tbl_bitwidth : int
            bitwidth for each bit-field
        type_bitwidth : int
            bitwith for bit-field base type
        tbl_struct_blk_num : int
            number of block calculated in bithwith of base type
        bf_blks : list
            a list of bit-field block lists, each bit-field block is a list of info for a specific bit-field
        bf_pos_info : list
            list of tup(upper position, bitwidth) for each bit-field
    """
    bf_col_min = sheet[tbl_struct_pos].column # column index
    bf_row_min = sheet[tbl_struct_pos].row # row index
    dim = sheet[tbl_bitwidth_pos].value

    if isinstance(dim, str):
        dim_tup = parse_dim_info(dim, 2, sheet, tbl_bitwidth_pos)
        tbl_bitwidth = dim_tup[0] * dim_tup[1]
    elif isinstance(dim, int):
        tbl_bitwidth = dim
    else:
        p_red2("value is wrong at cell", tbl_bitwidth_pos)
        sys.exit()
    type_bitwidth = sheet.cell(column=bf_col_min, row = bf_row_min - 1).value + 1 # basic data type size in bits
    if type_bitwidth == None or not isinstance(type_bitwidth, int):
        p_red2("value is wrong at cell", "{}{}".format(get_column_letter(bf_col_min), bf_row_min - 1))
        sys.exit()
    tbl_struct_blk_num = tbl_bitwidth // type_bitwidth

    # check single entry bitwidth info
    if ((sheet.max_row - tbl_struct_blk_num * bf_info_depth > 4)
            and (sheet.cell(column=2, row=bf_row_min + tbl_struct_blk_num * bf_info_depth ).value == None)
            or (sheet.max_row - tbl_struct_blk_num * bf_info_depth < 4)):
        p_purple_red6('num of field blocks', tbl_struct_blk_num, 'is wrong in sheet', sheet.title, 'at cell', tbl_bitwidth_pos)
        sys.exit()

    bf_blks=[] # [[bit_fields], [...] ]
    bf_pos_info = []
    bf_names = [] # for checking duplication
    bf_max_len = 0 # for bitfield formting
    for blk_num in range(tbl_struct_blk_num):
        bitwidth_check = type_bitwidth
        itr_col = bf_col_min
        itr_row = bf_row_min + blk_num * bf_info_depth
        bf_blk = []
        bf_pos = []
        bitwidth = 0
        pos = type_bitwidth - 1
        while bitwidth_check <= type_bitwidth and bitwidth_check > 0:
            bf =[] # bit_fields: ['name', 'description', 'unit', 'granularity', 'min', 'max', 'sign', 'bit width', 'upper position']
            bitwidth = sheet.cell(row = itr_row + 7, column = itr_col).value # bitwidth
            if bitwidth == None:
                p_red4('missing/wrong bit width around cell', "{}{}".format(get_column_letter(itr_col), itr_row + 7), 'in sheet', sheet.title)
                sys.exit()
            if  bitwidth <= 0:
                p_red4('wrong bit width at cell', "{}{}".format(get_column_letter(itr_col), itr_row + 7), 'in sheet', sheet.title)
                sys.exit()
            val = sheet.cell(row = itr_row, column = itr_col).value # name
            if val == None or val.lower() == 'res':
                if bitwidth == 1:
                    val = 'res_' + str(blk_num * type_bitwidth + pos)
                else:
                    val = 'res_' + str(blk_num * type_bitwidth + pos) + '_' + str(blk_num * type_bitwidth + pos - bitwidth + 1)
            val = val.lower() # XXX forcing bitfield name to lower case
            len_name = len(val)
            bf_names.append(val)
            if len(bf_names) > len(set(bf_names)):
                p_red4('Duplicated bit field name', val, 'in sheet', sheet.title)
                sys.exit()
            bf.append(val)
            val = sheet.cell(row = itr_row + 1, column = itr_col).value # description
            if val == None:
                val = 'TODO: missing description!!'
            bf.append(val)
            bf.append(sheet.cell(row = itr_row + 2, column = itr_col).value) # unit
            bf.append(sheet.cell(row = itr_row + 3, column = itr_col).value) # granularity
            val = sheet.cell(row = itr_row + 4, column = itr_col).value # min
            if val == None:
                val = 0
            elif isinstance(val, str) and val.startswith('0x'):
                val = int(str(val), 0)
            bf.append(val)
            #bf.append(sheet.cell(row = itr_row + 4, column = itr_col).value) # min
            val = sheet.cell(row = itr_row + 5, column = itr_col).value # min
            if val == None:
                val = 0
            elif isinstance(val, str) and val.startswith('0x'):
                val = int(str(val), 0)
            bf.append(val)
            #bf.append(sheet.cell(row = itr_row + 5, column = itr_col).value) # max
            val = sheet.cell(row = itr_row + 6, column = itr_col).value # sign
            if val == 'signed' or val == 'unsigned' or val == 'float' or val == 'double':
                bf.append(val)
            elif val != None and isinstance(val, str) and val[:2].lower()=='e_': # enum
                e_val = (val + '_t' , val)[val.lower()[-2:] == '_t'].lower()
                bf.append(e_val)
                len_name = len(e_val) # for adjusting output for enum case
                if val.lower()[-2:] != '_t': # only add enum when the given enum type name not ending with '_t'
                    if (e_val) not in enum_dic.keys():
                        p_green4("Generating new enum type name from", val, "as", e_val)
                        enum_dic[e_val] = [bf_names[-1].upper()]
                    else:
                        enum_dic[e_val].append(bf_names[-1].upper())
                else:
                    p_purple3('Please remember to add header file for your enum', val.lower(), 'with --hdr option')
            elif val == None:
                bf.append('unsigned')
                p_red4('bit field sign is missing at cell', "{}{}".format(get_column_letter(itr_col), itr_row + 6), 'in sheet', sheet.title)
                p_red2('set it to ', 'unsigned')
            else:
                p_red4('bit field sign is wrong at cell', "{}{}".format(get_column_letter(itr_col), itr_row + 6), 'in sheet', sheet.title)
                sys.exit()
            len_type = len(val)
            if (len_name + len_type) > bf_max_len:
                bf_max_len = len_name + len_type
            bf.append(bitwidth)
            bf.append(pos)
            bf_blk.append(bf)
            bf_pos.append((pos, bitwidth)) # also save bit fields upper position info based on block size
            bitwidth_check -= bitwidth
            if bitwidth_check < 0:
                p_red4('total bitwidth is wrong at row', itr_row + 7, 'in sheet', sheet.title)
                sys.exit()

            pos -= bitwidth
            if bitwidth_check >= 0:
                for c in range(1,bitwidth):
                    test = sheet.cell(row = itr_row + 7, column = itr_col + c).value # bitwidth
                    if test != None:
                        p_red4('larger than expected bit width given at cell', "{}{}".format(get_column_letter(itr_col), itr_row + 7), 'in sheet', sheet.title)
                        sys.exit()
                itr_col += bitwidth # move to next bit field
        bf_blk.reverse()
        bf_blks.append(bf_blk)
        bf_pos.reverse()
        bf_pos_info.append(bf_pos)
    return (tbl_bitwidth, type_bitwidth, tbl_struct_blk_num, bf_blks, bf_pos_info, bf_max_len)

def build_sub_tbl_info(sheet, tbl_bitwidth_tup, tbl_dim_tup, tbl_struct_pos, bf_info_depth, bf_pos_info):
    """Build sub-tbl info for each table

    Constructing sub-table info for each table

    Parameters
    ----------
    sheet : Worksheet
        the worksheet which contains all bitfield info for a table
    tbl_bitwidth_tup : tup
        a tup with bitwidth info for each table
    tbl_dim_tup : tup
        a tup with dimision info for each table
    tbl_struct_pos : str
        cell position info for table struct
    bf_info_depth : int
        number of info entries for each bit-field, this is a constant
    bf_pos_info : list
        list of tup(upper position, bitwidth) for each bit-field

    Returns
    ----------
    sub_tbl_dic : dict
        a dictionary, with subtable name as key, and a tup (L, R, tup_of_hex_value_of_a_sub_table) as value
    """
    table_pos_info = get_table_pos_info()
    sub_table_dim_col = table_pos_info[5]
    bf_col_min = sheet[tbl_struct_pos].column# column index
    bf_row_min = sheet[tbl_struct_pos].row # row index

    type_bitwidth = tbl_bitwidth_tup[1]
    tbl_struct_blk_num = tbl_bitwidth_tup[2]
    pos_info = [] # bf_pos_info[0],  bf_pos_info[1], ...
    for blk_num in range(tbl_struct_blk_num):
        pos_info += bf_pos_info[blk_num]

    # tbl_dim_tup = (A, B)
    # tbl_bitwidth_tup = (tbl_bitwidth, type_bitwidth, tbl_struct_blk_num, bf_max_len)
    current_sub_tbl_row = bf_row_min + bf_info_depth * tbl_struct_blk_num # first subtable row
    sub_tbl_dim = sheet.cell(row = current_sub_tbl_row, column = sub_table_dim_col).value
    sub_tbl_dic = {}
    a = tbl_dim_tup[0] # A
    while sub_tbl_dim != None and a > 0: # continue if not empty cell for dim cell
        sub_tbl_dim = parse_dim_info(sub_tbl_dim, 2, sheet, "{}{}".format(get_column_letter(sub_table_dim_col), current_sub_tbl_row))
        sub_tbl_visable_data_num = sub_tbl_dim[0] # L
        sub_tbl_name = sheet.cell(row = current_sub_tbl_row, column = bf_col_min).value
        if sub_tbl_name == None:
            p_red4('Missing sub table name at row', current_sub_tbl_row, 'in sheet', sheet.title)
            sys.exit()
        else:
            sub_tbl_name = sub_tbl_name.lower() # XXX forcing sub_tbl_name to lower case
        L, R, sig  = sub_tbl_dim # (L, R, *+_)
        B = tbl_dim_tup[1] # (A, B)
        if sig == '*':
            if L >= B:
                L = B
                R = 1
            elif L * R > B:
                R = B // L
            else:
                pass
        else:
            L = (min(L-R, B), min(L, B))[sig == '+']
            R = 1 # always set to 1 instead of 0 for easy usage when do _data.h generation
        # now L holds the valid length for initializaing
        sub_tbl_bf_info = ()# contains bit field values for each instance
        for entry_idx in range(L):
            bf_val = ()
            # go over each block (as one base_type in size) in this entry:
            for blk_num in range(tbl_struct_blk_num):
                #bf_pos_info: (low_pos, width)
                for itr_row in range(current_sub_tbl_row + 1 + blk_num + entry_idx * tbl_struct_blk_num, \
                        current_sub_tbl_row + 1 + (blk_num + 1) + entry_idx * tbl_struct_blk_num, tbl_struct_blk_num):
                    for pos in bf_pos_info[blk_num]:
                        val = sheet.cell(row = itr_row, column = type_bitwidth - pos[0] + bf_col_min - 1).value
                        #bf_val.append(val)
                        bf_val += (val,)
            sub_tbl_bf_info += (bf_val,)
        sub_tbl_hex = () # contains hex values for each instance
        for bf_val in sub_tbl_bf_info:
            combined_val = 0
            for v, (p, w) in zip(bf_val, pos_info): # v: bit_field_value, w: bit_field_width, p: bit_field_pos
                try:
                    combined_val |= (v & (2**w - 1)) << (p + 1 - w)
                except TypeError:
                    p_purple3('seems bitfield definition changed in sheet', sheet.title, 'but subtable not updated accordingly?')
                    sys.exit()
                if p == type_bitwidth - 1:
                    #print("{0:{1}b}".format(combined_val, type_bitwidth))
                    sub_tbl_hex += (hex(combined_val),)
                    combined_val = 0
        sub_tbl_dic[sub_tbl_name] = (L, R, sub_tbl_hex, sub_tbl_bf_info)
        current_sub_tbl_row += sub_tbl_visable_data_num * tbl_struct_blk_num + 1
        sub_tbl_dim = sheet.cell(row = current_sub_tbl_row, column = sub_table_dim_col).value
        a -= 1
    # sub_tbl_dic = {sub_tbl_name:(L, R, sub_tbl_hex, sub_tbl_bf_info), ...}
    return sub_tbl_dic # {sub_tbl_name:(L, R, sub_tbl_hex, sub_tbl_bf_info), ...}

def process_workbook(workbook, worksheets, table_dim_dic): # {tbl_name:(A, B), ...}
    """process workbook to extract table info

    this routine processes the workbook to extract all table info and orgnize in a dictionary

    Parameters
    ----------
    workbook : Workbook
        workbook of Excel File, contains all worksheet
    worksheets : set
        set of Worksheet, each worksheet provides info for a individual table
    table_dim_dic : dict
        a dimision dictionary: tbl_name as key, tup (sub_tbl_num_max, sub_tbl_entry_num_max) as value
        sub_tbl_num_max: constant, subtable first dimision
        sub_tbl_entry_num_max: constant, subtable second dimision

    Returns
    ----------
    tbl_dic : dict
        a dictionary with tbl_name as key, tup (tbl_des, bf_blks, sub_tbl_dic, tbl_bitwidth_tup, align) as value
        tbl_des: table description
        bf_blks: bit-field block list, each bit-field block is a list of info for a specific bit-field
        sub_tbl_dic: a dictionary with info about a subtable
        tbl_bitwidth_tup: a tup with bitwidth info for each table
        align: specifying the alignment attribute of this table
    enum_dic : dict
        a dictionary with enum type as keys, enum elements as values
    """
    table_pos_info = get_table_pos_info()
    tbl_name_pos = table_pos_info[0]
    tbl_des_pos = table_pos_info[1]
    tbl_bitwidth_pos = table_pos_info[2]
    tbl_struct_pos = table_pos_info[3]
    bf_info_depth = table_pos_info[4]
    tbl_alignment_pos = table_pos_info[6]
    tbl_names = {} # tbl_name:sheet.title
    tbl_dic = {}
    enum_dic = {} # enum dic: type:{element,...}
    for ws in worksheets:
        sheet = workbook[ws]
        keyinfo_check(sheet, tbl_name_pos, tbl_bitwidth_pos, tbl_struct_pos)
        tbl_name = sheet[tbl_name_pos].value
        tbl_name = tbl_name.lower() # XXX forcing table name to low case
        tbl_des = sheet[tbl_des_pos].value
        if tbl_name not in tbl_names.keys():
            tbl_names[tbl_name] = sheet.title
        else:
            p_purple_red6('Duplicated table name', tbl_name, 'in sheet', tbl_names[tbl_name], 'and', sheet.title)
            sys.exit()
        tbl_bitwidth, type_bitwidth, tbl_struct_blk_num, bf_blks, bf_pos_info, bf_max_len = build_bf_struct(sheet, tbl_bitwidth_pos, tbl_struct_pos, bf_info_depth, enum_dic)
        tbl_bitwidth_tup = (tbl_bitwidth, type_bitwidth, tbl_struct_blk_num, bf_max_len)
        if table_dim_dic != {}:
            tbl_dim_tup = table_dim_dic[tbl_name]
            sub_tbl_dic = build_sub_tbl_info(sheet, tbl_bitwidth_tup, tbl_dim_tup, tbl_struct_pos, bf_info_depth, bf_pos_info)
        else:
            sub_tbl_dic = {}
        # sub_tbl_dic = sub_tbl_name:{L, R, sub_tbl_hex, sub_tbl_bf_info}, ...}}
        tbl_dic[tbl_name] = (tbl_des, bf_blks, sub_tbl_dic, tbl_bitwidth_tup,  sheet[tbl_alignment_pos].value)
        #pp.pprint(tbl_dic)
    return (tbl_dic, enum_dic)
